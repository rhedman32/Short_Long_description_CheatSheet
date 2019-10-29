import openpyxl
import csv
import datetime
import os
from tkinter import *
from tkinter import ttk
from tkinter import messagebox

wb = openpyxl.Workbook()
sheet = wb.active

class Short_Long:

    def __init__(self, master):
        #Styling UI
        master.title('Short & Long Description Cheat Sheet')
        master.iconbitmap('Masonite.ico')
        # master.resizable(False, False)
        master.configure(background='#DAF7A6')

        self.style = ttk.Style()
        self.style.configure('TFrame', background='#DAF7A6')
        self.style.configure('TLabel', background='#DAF7A6', font=('Aerial', 10, 'bold'))
        self.style.configure('Header.TLabel', foreground='white', background='#90c74c', font=('Aerial', 18,'bold'))

        #Frame Header Setup
        self.frame_header = ttk.Frame(master)
        self.frame_header.pack()
        # self.frame_header.columnconfigure(0, weight=1)
        
        self.logo = PhotoImage(file = 'Masonite PP Banner.png')
        ttk.Label(self.frame_header, image = self.logo).grid(row=0, column=0, columnspan=4)
        ttk.Label(self.frame_header, text = 'Short & Long Description', style='Header.TLabel').grid(row=0, column=0, rowspan=2, columnspan=4)
        ttk.Label(self.frame_header, text = 'Long', foreground='#90c74c').grid(row=2, column=1, sticky='e')
        
        self.longer = BooleanVar()
        self.entry_long = ttk.Checkbutton(self.frame_header, variable=self.longer, onvalue=True, offvalue=False)
        self.entry_long.grid(row=2, column=2, sticky='w')

        #Frame Content Setup
        self.frame_content = ttk.Frame(master)
        self.frame_content.pack(expand=True)
        
        #Labels for Questions
        ttk.Label(self.frame_content, text = 'Backend Question Name:').grid(row=0, column=0, pady=3, sticky='e')
        ttk.Label(self.frame_content, text = 'Backend Group Question Name:').grid(row=1, column=0, pady=3, sticky='e')
        ttk.Label(self.frame_content, text = 'Order of Description:').grid(row=2, column=0, pady=3, sticky='e')
        ttk.Label(self.frame_content, text = 'Display Format:').grid(row=3, column=0, pady=3, sticky='e')
        ttk.Label(self.frame_content, text = 'Ignore Blank Answer:').grid(row=4, column=0, pady=3, sticky='e')
        ttk.Label(self.frame_content, text = 'Backend Answer to Hide:').grid(row=5, column=0, pady=3, sticky='e')
        ttk.Label(self.frame_content, text = 'Type in a Prefix:').grid(row=6, column=0, pady=3, sticky='e')
        ttk.Label(self.frame_content, text = 'Type in a Suffix:').grid(row=7, column=0, pady=3, sticky='e')
        ttk.Label(self.frame_content, text = 'Display Answer Override:').grid(row=8, column=0, pady=3, sticky='e')
        ttk.Label(self.frame_content, text = 'Overridden Answer Backend Name:').grid(row=9, column=0, pady=3, sticky='e')
        ttk.Label(self.frame_content, text = 'Fraction or Decimal:').grid(row=10, column=0, pady=3, sticky='e')
        
        #Entry Values for Questions
        self.DFormat = StringVar()
        self.IgnoreBA = BooleanVar()
        self.Fraction = StringVar()
        self.row = 0
        self.order=StringVar()
    
        self.entry_question = ttk.Entry(self.frame_content, width = 24)
        self.entry_group_question = ttk.Entry(self.frame_content, width = 24)
        self.entry_order = ttk.Entry(self.frame_content, width = 24, textvariable=self.order)
        self.entry_display = ttk.Combobox(self.frame_content, textvariable=self.DFormat)
        self.entry_display.config(values=('Answer Only', 'Name Only', 'Name and Answer'), width = 21)
        self.entry_ignore = ttk.Checkbutton(self.frame_content)
        self.entry_ignore.config(variable=self.IgnoreBA, onvalue=True, offvalue=False)
        self.entry_hide = ttk.Entry(self.frame_content, width = 24)
        self.hide_plus = ttk.Button(self.frame_content, text = '+',  command=self.additionalAnswers)
        self.entry_prefix = ttk.Entry(self.frame_content, width = 24)
        self.entry_suffix = ttk.Entry(self.frame_content, width = 24)
        self.entry_change_answer = ttk.Entry(self.frame_content, width = 24)
        self.entry_change_backend = ttk.Entry(self.frame_content, width = 24)
        self.change_plus = ttk.Button(self.frame_content, text = '+', command=self.additionalChangedAnswers)
        self.entry_decimal = ttk.Combobox(self.frame_content, textvariable=self.Fraction)
        self.entry_decimal.config(values=('Decimal', 'Fraction'), width = 21)

        self.entry_question.grid(row=0, column=1, pady=5)
        self.entry_group_question.grid(row=1, column=1, pady=5)
        self.entry_order.grid(row=2, column=1, pady=5)
        self.entry_display.grid(row=3, column=1, pady=5)
        self.entry_ignore.grid(row=4, column=1, sticky='w', pady=5)
        self.entry_hide.grid(row=5, column=1, pady=5)
        self.hide_plus.grid(row=5, column=2, pady=5)
        self.entry_prefix.grid(row=6, column=1, pady=5)
        self.entry_suffix.grid(row=7, column=1, pady=5)
        self.entry_change_answer.grid(row=8, column=1, pady=5)
        self.entry_change_backend.grid(row=9, column=1, pady=5)
        self.change_plus.grid(row=9, column=2, pady=5)
        self.entry_decimal.grid(row=10, column=1, pady=5)

        self.list_view = ttk.Treeview(self.frame_content)
        self.list_view.grid(row=0, column=2, rowspan=5, columnspan=2, padx=10, pady=5)
        self.list_view.heading('#0',text='Questions Inserted')

        self.adding = ttk.Button(self.frame_content, text = 'Add', command=self.add)
        self.adding.grid(row=12, column=1, pady=3)
        ttk.Button(self.frame_content, text = 'Spit Out', command=self.SpitOut).grid(row=12, column=2, pady=3)
        self.adding.state(['disabled'])

        def buttonFunction(*args):
            if self.entry_question.get() and self.entry_group_question.get() and self.entry_order.get():
                self.adding.state(['!disabled'])
        
        self.order.trace('w', buttonFunction)

    def LastColumn(self, D):
        Desc = str(D)
        while len(Desc) != 5:
            Desc = '0' + Desc
        return Desc

    def DescriptionOrder(self, att2, att3, attValue1):
        DO = 'Default,'+ att2 + ',' + att3 + ',' + attValue1 + ',' + self.LastColumn(attValue1)
        return DO

    def DisplayFormat(self, att3, displayValue, desc):
        DF = 'Default,DisplayFormat,' + att3 + ',' + displayValue + ',' + self.LastColumn(desc)
        return DF

    def IgnoringAnswer(self, att1, att3, boolValue, desc):
        IA = att1 + ',IgnoreBlankAnswers,' + att3 + ',' + boolValue + ',' + self.LastColumn(desc)
        return IA

    def AnswerVisibility(self, att1, att3, att3b, desc):
        AV = att1 + ',ItemVisible,' + att3 + ':' + att3b + ',False' + ',' + self.LastColumn(desc)
        return AV

    def Prefix(self, att1, att3, text, desc):
        PRE = att1 + ',QuestionAnswerPrefix,' + att3 + ',' + text + ',' + self.LastColumn(desc)
        return PRE

    def Suffix(self, att1, att3, text, desc):
        SUF = att1 + ',QuestionAnswerSuffix,' + att3 + ',' + text + ',' + self.LastColumn(desc)
        return SUF

    def ChangeDisplayName(self, att1, att3, att3b, text, desc):
        CDN = att1 + ',QuestionAnswerName,' + att3 + ':' + att3b + ',' + text + ',' + self.LastColumn(desc)
        return CDN

    def DisplayFractionDecimal(self, att1, att3, boolValue, desc):
        DFD = att1 + ',ReportAsDecimal,' + att3 + ',' + boolValue + ',' + self.LastColumn(desc)
        return DFD

    def NewEndLine(self, att1, att3, boolValue, desc):
        NEL = att1 + ',EndsLine,' + att3 + ',' + boolValue + ',' + self.LastColumn(desc)
        return NEL

    def additionalAnswers(self):
        Visibility = self.entry_hide.get()
        A3 = self.entry_question.get()
        A1 = self.entry_group_question.get()
        AV1 = self.entry_order.get()
        if Visibility:
            self.row+=1
            VisibleText = self.AnswerVisibility(A1, A3, Visibility, AV1)
            text = VisibleText.split(',')
            for i, cell in enumerate(text):
                sheet.cell(row=self.row, column=i+1).value = cell
        self.entry_hide.delete(0, 'end')
    
    def additionalChangedAnswers(self):
        ChangeName = self.entry_change_answer.get()
        if ChangeName:
            AnswerName = self.entry_change_backend.get()
        A3 = self.entry_question.get()
        A1 = self.entry_group_question.get()
        AV1 = self.entry_order.get()
        if ChangeName:
            self.row+=1
            CN = self.ChangeDisplayName(A1, A3, AnswerName, ChangeName, AV1)
            text = CN.split(',')
            for i, cell in enumerate(text):
                sheet.cell(row=self.row, column=i+1).value = cell
        self.entry_change_answer.delete(0, 'end')
        self.entry_change_backend.delete(0, 'end')

    #Clears the entry fields and 
    def add(self):
        A3 = self.entry_question.get()
        A1 = self.entry_group_question.get()
        AV1 = self.entry_order.get()
        Format = self.DFormat.get()
        Ignore = self.IgnoreBA.get()
        NewLine = self.longer.get()
        if Ignore:
            Ignore = 'True'
        else:
            Ignore = ''
        Visibility = self.entry_hide.get()
        PrefixText = self.entry_prefix.get()
        SuffixText = self.entry_suffix.get()
        ChangeName = self.entry_change_answer.get()
        if ChangeName:
            AnswerName = self.entry_change_backend.get()
        FractionDecimal = str(self.Fraction)
        if FractionDecimal == 'Decimal':
            FractionDecimal = 'True'
        elif FractionDecimal == 'Fraction':
            FractionDecimal = 'False'
        else:
            FractionDecimal = ''
        if NewLine:
            NewLine = 'True'
        else:
            NewLine = ''

        # initRow = self.row+1

        if AV1:
            # if self.row == 0:
            #     sheet = wb.active
            self.row+=1
            DO1 = self.DescriptionOrder(A1, A3, AV1)
            text = DO1.split(',')
            for i, cell in enumerate(text):
                print(i)
                print(cell)
                print(self.row)
                sheet.cell(row=self.row, column=i+1).value = cell
                # sheet[col[i]+str(self.row)] = cell

        if Format:
            self.row+=1
            DF1 = self.DisplayFormat(A3, Format, AV1)
            text = DF1.split(',')
            for i, cell in enumerate(text):
                sheet.cell(row=self.row, column=i+1).value = cell

        if Ignore:
            self.row+=1
            IgnoreText = self.IgnoringAnswer(A1, A3, Ignore, AV1)
            text = IgnoreText.split(',')
            for i, cell in enumerate(text):
                sheet.cell(row=self.row, column=i+1).value = cell
        
        if Visibility:
            self.row+=1
            VisibleText = self.AnswerVisibility(A1, A3, Visibility, AV1)
            text = VisibleText.split(',')
            for i, cell in enumerate(text):
                sheet.cell(row=self.row, column=i+1).value = cell

        if PrefixText:
            self.row+=1
            PFT = self.Prefix(A1, A3, PrefixText, AV1)
            text = PFT.split(',')
            for i, cell in enumerate(text):
                sheet.cell(row=self.row, column=i+1).value = cell
        
        if SuffixText:
            self.row+=1
            SFT = self.Suffix(A1, A3, SuffixText, AV1)
            text = SFT.split(',')
            for i, cell in enumerate(text):
                sheet.cell(row=self.row, column=i+1).value = cell

        if ChangeName:
            self.row+=1
            CN = self.ChangeDisplayName(A1, A3, AnswerName, ChangeName, AV1)
            text = CN.split(',')
            for i, cell in enumerate(text):
                sheet.cell(row=self.row, column=i+1).value = cell
            
        if FractionDecimal:
            self.row+=1
            FD = self.DisplayFractionDecimal(A1, A3, FractionDecimal, AV1)
            text = FD.split(',')
            for i, cell in enumerate(text):
                sheet.cell(row=self.row, column=i+1).value = cell
        
        if NewLine:
            self.row+=1
            NL = self.NewEndLine(A1, A3, NewLine, AV1)
            text = NL.split(',')
            for i, cell in enumerate(text):
                sheet.cell(row=self.row, column=i+1).value = cell
        
        # for x in range(initRow, self.row+1):
        #     print('***')
        #     Desc = str(AV1)
        #     while len(Desc) != 5:
        #         Desc = '0' + Desc
        #     sheet['E'+ str(x)] = Desc
        
        self.list_view.insert('', AV1, A3, text=AV1+' - '+A3)

        self.entry_question.delete(0, 'end')
        self.entry_group_question.delete(0, 'end')
        self.entry_order.delete(0, 'end')
        self.entry_display.delete(0, 'end')
        if Ignore == 'True':
            self.IgnoreBA.set(False)
        self.entry_hide.delete(0, 'end')
        self.entry_prefix.delete(0, 'end')
        self.entry_suffix.delete(0, 'end')
        self.entry_change_answer.delete(0, 'end')
        self.entry_change_backend.delete(0, 'end')
        self.entry_decimal.delete(0, 'end')
        self.adding.state(['disabled'])

    #Excel file is generated and clears the treeview pane
    def SpitOut(self):
        self.row = 0
        if self.list_view.get_children():
            wb.save(os.getcwd() + '\\' + '{:%m-%d-%Y-%H_%M_%S} '.format(datetime.datetime.now()) + '_Short_Description.csv')
            messagebox.showinfo(title='Your File is Ready!', message='An excel file has been generated check \n'+os.getcwd())
        for x in self.list_view.get_children():
            self.list_view.delete(x)
        for i in range(1,sheet.max_row+1):
            for j in range(1,sheet.max_column+1):
                sheet.cell(row=i, column=j).value = ''

def main():
    root = Tk()
    ShortAndLong = Short_Long(root)
    print(root.winfo_height())
    print(root.winfo_width())
    root.mainloop()
    

if __name__ == "__main__": main()



